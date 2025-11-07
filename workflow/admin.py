
from workflow.models import PLMWindchillMockdata, Rule , UdiFiaWorkflow

# admin.py
import io
from django.contrib import admin, messages
from django.urls import path, reverse
from django.shortcuts import redirect, render
from django.utils.html import format_html
from django.db import transaction
from openpyxl import load_workbook

from .models import Rule
from .forms import RuleSchemaUploadForm

# Mapping from Excel header to Rule model field name.
# Update keys to match exact header text in your Excel file.
EXCEL_TO_MODEL_FIELD = {
    "HEALTH AUTHORITY": "health_authority",
    "UDI REGULATION": "udi_regulation",
    "CATEGORY": "category",
    "DATA PROPERTY": "data_property",
    "DATA ATTRIBUTE (HA Field Name)": "data_attribute_ha_field_name",
    "GUDE FIELD NAME": "gude_field_name",
    "JNJ UDI Data Element": "jnj_udi_data_element",
    "GUDE FIELD NUMBER #": "gude_field_number",
    "BUDI ATTRIBUTE (EUDAMED-only)": "budi_attribute_eudamed_only",
    "GS1 GTIN TRIGGER (100782299 APPENDIX B)": "gs1_gtin_trigger_100782299_appendix_b",
    "HEALTH AUTHORITY GTIN TRIGGER": "health_authority_gtin_trigger",
    "JJMT USE DIRECTIVE (Rules-based / Optional Use)": "jjmt_use_directive",
    "MANDATORY FIELD IN DATABASE (Yes/No)": "mandatory_field_in_database",
    "FIELD TYPE": "field_type",
    "ADD (Yes/No)": "add_flag",
    "EDIT (Yes/No)": "edit_flag",
    "DELETE (Yes/No)": "delete_flag",
    "CHANGE CONDITION or SCENARIOS": "change_condition_or_scenarios",
    "ADDITIONAL CHANGE REQUEST REQUIREMENTS": "additional_change_request_requirements",
    "DRI Comments": "dri_comments",
    "GTIN OUTCOME ACTION": "gtin_outcome_action",
    "DATA SOURCE OUTCOME ACTION": "data_source_outcome_action",
    # Add any other header->field mappings if needed
}

@admin.register(Rule)
class RuleAdmin(admin.ModelAdmin):
    list_display = [
        'pk',
        'health_authority',
        'udi_regulation',
        'category',
        'data_property',
        'gude_field_name',
        'jnj_udi_data_element',
        'gude_field_number',
        'field_type',
        'mandatory_field_in_database',
        'add_flag',
        'edit_flag',
        'delete_flag',
        'gtin_outcome_action',
        'data_source_outcome_action',
    ]
    list_display_links = ['pk']
    # (your other admin config here...)

    change_list_template = "admin/rules_change_list_with_upload.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('upload-schema/', self.admin_site.admin_view(self.upload_schema_view), name='rule_upload_schema'),
        ]
        return my_urls + urls

    def upload_schema_view(self, request):
        """
        Custom admin view to handle Excel uploads.
        Deletes existing Rule rows and creates new ones from the uploaded sheet.
        """
        # compute changelist url dynamically for redirects/templates
        changelist_url = reverse("admin:workflow_rule_changelist")

        if request.method == "POST":
            form = RuleSchemaUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data['excel_file']
                try:
                    in_memory = excel_file.read()
                    wb = load_workbook(filename=io.BytesIO(in_memory), data_only=True)
                    ws = wb.active  # Use the first sheet

                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        messages.error(request, "Uploaded file is empty.")
                        return redirect(changelist_url)
                    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

                    # Build header->model_field mapping (for headers present in the file)
                    header_to_field = {}
                    for idx, header in enumerate(headers):
                        if header in EXCEL_TO_MODEL_FIELD:
                            header_to_field[idx] = EXCEL_TO_MODEL_FIELD[header]
                        else:
                            # tolerant match fallback
                            normalized = ''.join(ch for ch in header.upper() if ch.isalnum())
                            found = None
                            for k, v in EXCEL_TO_MODEL_FIELD.items():
                                kn = ''.join(ch for ch in k.upper() if ch.isalnum())
                                if kn == normalized:
                                    found = v
                                    break
                            if found:
                                header_to_field[idx] = found
                            else:
                                header_to_field[idx] = None

                    data_rows = rows[1:]
                    created_count = 0

                    with transaction.atomic():
                        Rule.objects.all().delete()

                        for row in data_rows:
                            if all(cell is None for cell in row):
                                continue
                            obj_kwargs = {}
                            for col_idx, cell_value in enumerate(row):
                                model_field = header_to_field.get(col_idx)
                                if not model_field:
                                    continue
                                if isinstance(cell_value, str):
                                    val = cell_value.strip()
                                else:
                                    val = cell_value
                                obj_kwargs[model_field] = val
                            Rule.objects.create(**obj_kwargs)
                            created_count += 1

                    messages.success(request, f"Successfully imported {created_count} rules. Existing data was cleared before import.")
                    return redirect(changelist_url)
                except Exception as e:
                    messages.error(request, f"Error processing file: {e}")
                    return redirect(changelist_url)
        else:
            form = RuleSchemaUploadForm()

        context = dict(
            self.admin_site.each_context(request),
            title="Upload Rule schema (.xlsx)",
            form=form,
            changelist_url=changelist_url,  # pass URL to template
        )
        return render(request, "admin/rule_upload_form.html", context)

    def changelist_view(self, request, extra_context=None):
        extra_context = (extra_context or {})
        # pass upload URL name to changelist template (if you use it)
        extra_context['upload_url_name'] = 'admin:rule_upload_schema'
        return super().changelist_view(request, extra_context=extra_context)


admin.site.register( UdiFiaWorkflow)
admin.site.register( PLMWindchillMockdata)